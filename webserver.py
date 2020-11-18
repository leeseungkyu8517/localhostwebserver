# WebServer.py

from matplotlib import pyplot as plt
from matplotlib import animation
import numpy as np
import openpyxl
import datetime
from datetime import time
from tkinter import Tk

import sys

from PyQt5.QtWidgets import QApplication, QWidget
from tkinter import messagebox as msg




wb = openpyxl.load_workbook('test.xlsx')

# 현재 Active Sheet 얻기
ws = wb.active
# ws = wb.get_sheet_by_name("Sheet1")

for r in ws.rows:
    row_index = r[0].row  # 행 인덱스

from socket import *

serverPort = 8000   # server port
serverAddress = ""  # server address
serverSocket = socket(AF_INET, SOCK_STREAM)
serverSocket.bind((serverAddress, serverPort))
serverSocket.listen(1000)   # 1000 = number of unexpected connection
print("Ready to server")

countad=0
sumad=0
compsec = 61


class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('My First Application')
        self.move(300, 300)
        self.resize(400, 200)
        self.show()


while True:


    connectionSocket, address = serverSocket.accept()
    now = datetime.datetime.now()



    try:
        message = connectionSocket.recv(1024)   # buf size 1024
        print("Message : " + str(message))
        # exception when message is null
        if not message:
            continue
        filename = message.split()[1]
        f = open(filename[1:], "rb")
        outputData = f.read()

        # 200 OK
        header = "HTTP/1.1 200 OK\r\n\r\n"
        headerBytes = bytes(header, "UTF-8")
        connectionSocket.send(headerBytes)
        # when buf size is 1 can't parse html code
        for i in range(0, len(outputData), 1024):   # send buf size 1024
            connectionSocket.send(outputData[i:i + 1024])
        connectionSocket.send(b"\r\n\r\n")
        countad = countad + 1
        sumad=sumad+1
        print("host", "here")
        print(countad)



        if compsec != int(now.second):

            ws.cell(row=sumad, column=5).value = str(message)

            print(str(now.second))
            ws.cell(row=sumad, column=3).value = str(sumad)
            ws.cell(row=sumad, column=2).value = str(now)

            countad=0
            ws.cell(row=sumad, column=4).value = str(countad+1)
            ws.cell(row=sumad, column=1).value = str(now.year) + str(now.month) + str(now.day) + str(now.hour) + str(
                now.minute) + str(now.second) + str("host") + str(countad)
            compsec = now.second
        else :
            ws.cell(row=sumad, column=5).value = str(message)
            ws.cell(row=sumad, column=4).value = str(countad+1)
            print(str(now.second))
            ws.cell(row=sumad, column=3).value = str(sumad)
            ws.cell(row=sumad, column=2).value = str(now)
            ws.cell(row=sumad, column=1).value = str(now.year) + str(now.month) + str(now.day) + str(now.hour) + str(now.minute) + str(now.second) + str("host") + str(countad)

        # 엑셀 파일 저장
        wb.save("test.xlsx")
        wb.close()




        connectionSocket.close()


    except IOError:
        # 404 Not Found
        header = "HTTP/1.1 404 Not found\r\n"
        headerBytes = bytes(header, "UTF-8")
        connectionSocket.send(headerBytes)
        countad = countad + 1
        sumad = sumad + 1
        print("stranger", "here")
        print(countad)

        if compsec != int(now.second):

            ws.cell(row=sumad, column=5).value = str(message)

            print(str(now.second))
            ws.cell(row=sumad, column=3).value = str(sumad)
            ws.cell(row=sumad, column=2).value = str(now)

            countad = 0
            ws.cell(row=sumad, column=4).value = str(countad+1)
            ws.cell(row=sumad, column=1).value = str(now.year) + str(now.month) + str(now.day) + str(now.hour) + str(
                now.minute) + str(now.second) + str("stranger") + str(countad)
            compsec = now.second
        else:
            ws.cell(row=sumad, column=5).value = str(message)
            ws.cell(row=sumad, column=4).value = str(countad+1)
            print(str(now.second))
            ws.cell(row=sumad, column=3).value = str(sumad)
            ws.cell(row=sumad, column=2).value = str(now)
            ws.cell(row=sumad, column=1).value = str(now.year) + str(now.month) + str(now.day) + str(now.hour) + str(
                now.minute) + str(now.second) + str("stranger") + str(countad)

        # 엑셀 파일 저장
        wb.save("test.xlsx")
        wb.close()

        connectionSocket.close()















